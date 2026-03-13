"""
kpi_cfar.py
-----------
Populates the Column Free Area Ratio (CFAR) sheet in the KPI workbook.

Model structure:
    Grasshopper Model (root)
    ├── Columns   -> objects with properties: level, column_area
    ├── Slabs     -> objects with properties: level, slab_area
    └── Cores     -> objects with properties: level, core_area

Formula per level:
    cf_ratio = (slab_area - (column_area + core_area)) / slab_area

Output layout:
    Row 1  — KPI heading (merged, Light Magenta 2)
    Row 2  — Column headers (darker magenta)
    Row 3  — Entire Building total (bold, white)
    Row 4+ — One row per level, alternating white / light grey 2
"""

from collections import defaultdict
from openpyxl.utils import get_column_letter

from _collection_helper import get_collection_objects, get_prop, get_level
from excel_formatter import (
    style_kpi_heading,
    style_column_headers,
    style_total_row,
    style_data_row,
    set_column_widths,
    freeze_below_headers,
)

HEADERS = [
    "Level",
    "Slab Area (m²)",
    "Column Area (m²)",
    "Core Area (m²)",
    "CF Ratio",
]
NUM_COLS = len(HEADERS)

C_SLAB  = "B"
C_COL   = "C"
C_CORE  = "D"
C_RATIO = "E"


def write_cfar_sheet(ws, root):
    """Write the Column Free Area Ratio KPI data into worksheet ws."""

    # ── 1. Collect objects ───────────────────────────────────────────────
    slab_objects   = get_collection_objects(root, "Slabs")
    column_objects = get_collection_objects(root, "Columns")
    core_objects   = get_collection_objects(root, "Cores")

    # ── 2. Aggregate areas per level ─────────────────────────────────────
    slab_by_level   = defaultdict(float)
    column_by_level = defaultdict(float)
    core_by_level   = defaultdict(float)

    for obj in slab_objects:
        slab_by_level[get_level(obj)] += float(get_prop(obj, "slab_area") or 0)

    for obj in column_objects:
        column_by_level[get_level(obj)] += float(get_prop(obj, "column_area") or 0)

    for obj in core_objects:
        core_by_level[get_level(obj)] += float(get_prop(obj, "core_area") or 0)

    # ── 3. Sorted level list ─────────────────────────────────────────────
    all_levels = sorted(
        set(list(slab_by_level) + list(column_by_level) + list(core_by_level))
    )

    first_data_row = 4
    last_data_row  = 3 + len(all_levels)

    # ── 4. KPI heading (row 1) ───────────────────────────────────────────
    ws.cell(row=1, column=1, value="Column Free Area Ratio (CFAR)")
    style_kpi_heading(ws, row=1, num_cols=NUM_COLS, kpi="cfar")

    # ── 5. Column headers (row 2) ────────────────────────────────────────
    for col_i, header in enumerate(HEADERS, start=1):
        ws.cell(row=2, column=col_i, value=header)
    style_column_headers(ws, row=2, num_cols=NUM_COLS, kpi="cfar")

    # ── 6. Entire Building total (row 3) ─────────────────────────────────
    ws.cell(row=3, column=1, value="Entire Building")
    ws.cell(row=3, column=2, value=f"=SUM({C_SLAB}{first_data_row}:{C_SLAB}{last_data_row})")
    ws.cell(row=3, column=3, value=f"=SUM({C_COL}{first_data_row}:{C_COL}{last_data_row})")
    ws.cell(row=3, column=4, value=f"=SUM({C_CORE}{first_data_row}:{C_CORE}{last_data_row})")
    ws.cell(row=3, column=5, value=f"=({C_SLAB}3-({C_COL}3+{C_CORE}3))/{C_SLAB}3")
    style_total_row(ws, row=3, num_cols=NUM_COLS)

    # ── 7. Per-level rows (row 4+) ───────────────────────────────────────
    for i, level in enumerate(all_levels):
        r = i + first_data_row
        ws.cell(row=r, column=1, value=level)
        ws.cell(row=r, column=2, value=round(slab_by_level.get(level, 0.0),   4))
        ws.cell(row=r, column=3, value=round(column_by_level.get(level, 0.0), 4))
        ws.cell(row=r, column=4, value=round(core_by_level.get(level, 0.0),   4))
        ws.cell(row=r, column=5, value=f"=({C_SLAB}{r}-({C_COL}{r}+{C_CORE}{r}))/{C_SLAB}{r}")
        style_data_row(ws, row=r, num_cols=NUM_COLS, row_index=i)

    # ── 8. Number format + column widths + freeze ────────────────────────
    for r in range(3, last_data_row + 1):
        ws.cell(row=r, column=5).number_format = "0.00"

    set_column_widths(ws, [20, 18, 18, 16, 14])
    freeze_below_headers(ws, header_row=2)