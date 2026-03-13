"""
kpi_mui.py
----------
Populates the Material Usage Intensity (MUI) sheet in the KPI workbook.

Model structure:
    Grasshopper Model (root)
    ├── Columns   -> objects with properties: level, column_volume
    ├── Slabs     -> objects with properties: level, slab_volume, slab_area
    └── Cores     -> objects with properties: level, core_volume

Formula per level:
    MUI = (slab_volume + column_volume + core_volume) / slab_area

Output layout:
    Row 1  — KPI heading (merged, Light Red 2)
    Row 2  — Column headers (darker red)
    Row 3  — Entire Building total (bold, white)
    Row 4+ — One row per level, alternating white / light grey 2
"""

from collections import defaultdict
from openpyxl.utils import get_column_letter

from _collection_helper import get_collection_objects, get_prop, get_level
from excel_formatter import (
    style_kpi_heading,
    style_column_headers,
    style_total_row,
    style_data_row,
    set_column_widths,
    freeze_below_headers,
)

HEADERS = [
    "Level",
    "Slab Volume (m³)",
    "Column Volume (m³)",
    "Core Volume (m³)",
    "Total Volume (m³)",
    "Slab Area (m²)",
    "MUI (m³/m²)",
]
NUM_COLS = len(HEADERS)

C_SLAB_VOL  = "B"
C_COL_VOL   = "C"
C_CORE_VOL  = "D"
C_TOTAL_VOL = "E"
C_SLAB_AREA = "F"
C_MUI       = "G"


def write_mui_sheet(ws, root):
    """Write the Material Usage Intensity KPI data into worksheet ws."""

    # ── 1. Collect objects ───────────────────────────────────────────────
    slab_objects   = get_collection_objects(root, "Slabs")
    column_objects = get_collection_objects(root, "Columns")
    core_objects   = get_collection_objects(root, "Cores")

    # ── 2. Aggregate volumes and slab area per level ─────────────────────
    slab_vol_by_level  = defaultdict(float)
    col_vol_by_level   = defaultdict(float)
    core_vol_by_level  = defaultdict(float)
    slab_area_by_level = defaultdict(float)

    for obj in slab_objects:
        lvl = get_level(obj)
        slab_vol_by_level[lvl]  += float(get_prop(obj, "slab_volume")  or 0)
        slab_area_by_level[lvl] += float(get_prop(obj, "slab_area")    or 0)

    for obj in column_objects:
        col_vol_by_level[get_level(obj)] += float(get_prop(obj, "column_volume") or 0)

    for obj in core_objects:
        core_vol_by_level[get_level(obj)] += float(get_prop(obj, "core_volume") or 0)

    # ── 3. Sorted level list ─────────────────────────────────────────────
    all_levels = sorted(
        set(list(slab_vol_by_level) + list(col_vol_by_level) + list(core_vol_by_level))
    )

    first_data_row = 4
    last_data_row  = 3 + len(all_levels)

    # ── 4. KPI heading (row 1) ───────────────────────────────────────────
    ws.cell(row=1, column=1, value="Material Usage Intensity (MUI)")
    style_kpi_heading(ws, row=1, num_cols=NUM_COLS, kpi="mui")

    # ── 5. Column headers (row 2) ────────────────────────────────────────
    for col_i, header in enumerate(HEADERS, start=1):
        ws.cell(row=2, column=col_i, value=header)
    style_column_headers(ws, row=2, num_cols=NUM_COLS, kpi="mui")

    # ── 6. Entire Building total (row 3) ─────────────────────────────────
    ws.cell(row=3, column=1, value="Entire Building")
    ws.cell(row=3, column=2, value=f"=SUM({C_SLAB_VOL}{first_data_row}:{C_SLAB_VOL}{last_data_row})")
    ws.cell(row=3, column=3, value=f"=SUM({C_COL_VOL}{first_data_row}:{C_COL_VOL}{last_data_row})")
    ws.cell(row=3, column=4, value=f"=SUM({C_CORE_VOL}{first_data_row}:{C_CORE_VOL}{last_data_row})")
    ws.cell(row=3, column=5, value=f"={C_SLAB_VOL}3+{C_COL_VOL}3+{C_CORE_VOL}3")
    ws.cell(row=3, column=6, value=f"=SUM({C_SLAB_AREA}{first_data_row}:{C_SLAB_AREA}{last_data_row})")
    ws.cell(row=3, column=7, value=f"={C_TOTAL_VOL}3/{C_SLAB_AREA}3")
    style_total_row(ws, row=3, num_cols=NUM_COLS)

    # ── 7. Per-level rows (row 4+) ───────────────────────────────────────
    for i, level in enumerate(all_levels):
        r = i + first_data_row
        ws.cell(row=r, column=1, value=level)
        ws.cell(row=r, column=2, value=round(slab_vol_by_level.get(level,  0.0), 4))
        ws.cell(row=r, column=3, value=round(col_vol_by_level.get(level,   0.0), 4))
        ws.cell(row=r, column=4, value=round(core_vol_by_level.get(level,  0.0), 4))
        ws.cell(row=r, column=5, value=f"={C_SLAB_VOL}{r}+{C_COL_VOL}{r}+{C_CORE_VOL}{r}")
        ws.cell(row=r, column=6, value=round(slab_area_by_level.get(level, 0.0), 4))
        ws.cell(row=r, column=7, value=f"={C_TOTAL_VOL}{r}/{C_SLAB_AREA}{r}")
        style_data_row(ws, row=r, num_cols=NUM_COLS, row_index=i)

    # ── 8. Number format + column widths + freeze ────────────────────────
    for r in range(3, last_data_row + 1):
        ws.cell(row=r, column=7).number_format = "0.00"

    set_column_widths(ws, [20, 20, 20, 18, 18, 16, 16])
    freeze_below_headers(ws, header_row=2)