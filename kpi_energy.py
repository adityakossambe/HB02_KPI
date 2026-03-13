"""
kpi_energy.py
-------------
Populates the Energy & Radiation Performance sheet in the KPI workbook.

Model structure:
    Grasshopper Model (root)
    └── Facade -> mesh objects, each with:
            obj.id                    — unique mesh hash ID
            obj.properties.panel_id   — panel type label (e.g. T1, Q3)
            obj.properties.isr_value  — comma-separated string of 12 monthly
                                        radiation values (Jan -> Dec)

Output layout:
    Row 1  — KPI heading (merged, Light Yellow 2)
    Row 2  — Entire Building totals (bold, white)
    Row 3  — Column headers (darker yellow)
    Row 4+ — One row per mesh, alternating white / light grey 2
             Columns: Panel ID | Mesh ID | Jan..Dec | Annual
"""

from openpyxl.utils import get_column_letter

from _collection_helper import get_collection_objects, get_prop, id_sort_key
from excel_formatter import (
    style_kpi_heading,
    style_column_headers,
    style_total_row,
    style_data_row,
    set_column_widths,
    freeze_below_headers,
)

MONTHS = ["Jan", "Feb", "Mar", "Apr", "May", "Jun",
          "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]

COL_PANEL_ID = 1
COL_MESH_ID  = 2
COL_JAN      = 3
COL_ANNUAL   = 3 + len(MONTHS)
NUM_COLS     = 2 + len(MONTHS) + 1


def _parse_isr(isr_value) -> list:
    """Parse comma-separated isr_value string into a list of 12 floats."""
    try:
        values = [float(v.strip()) for v in str(isr_value).split(",")]
        return (values + [0.0] * 12)[:12]
    except (ValueError, AttributeError):
        return [0.0] * 12


def _annual_formula(row: int) -> str:
    return "=" + "+".join(
        get_column_letter(COL_JAN + m) + str(row) for m in range(12)
    )


def write_energy_sheet(ws, root):
    """Write the Energy & Radiation Performance KPI data into worksheet ws."""

    # ── 1. Collect facade objects ────────────────────────────────────────
    facade_objects = get_collection_objects(root, "Facade")

    # ── 2. Parse each mesh ───────────────────────────────────────────────
    rows = []
    for obj in facade_objects:
        mesh_id   = getattr(obj, "id", None) or str(id(obj))
        panel_id  = get_prop(obj, "panel_id") or "—"
        isr_value = get_prop(obj, "isr_value")
        monthly   = _parse_isr(isr_value)
        rows.append((str(panel_id), str(mesh_id), monthly))

    # Sort by panel_id so same types group together
    rows.sort(key=lambda x: id_sort_key(x[0]))

    first_data_row = 4
    last_data_row  = 3 + len(rows)

    # ── 3. KPI heading (row 1) ───────────────────────────────────────────
    ws.cell(row=1, column=1, value="Energy & Radiation Performance")
    style_kpi_heading(ws, row=1, num_cols=NUM_COLS, kpi="energy")

    # ── 4. Entire Building total (row 2) ─────────────────────────────────
    ws.cell(row=2, column=COL_PANEL_ID, value="Entire Building")
    ws.cell(row=2, column=COL_MESH_ID,  value="—")
    for m_i in range(len(MONTHS)):
        col_letter = get_column_letter(COL_JAN + m_i)
        ws.cell(row=2, column=COL_JAN + m_i,
                value=f"=SUM({col_letter}{first_data_row}:{col_letter}{last_data_row})")
    annual_letter = get_column_letter(COL_ANNUAL)
    ws.cell(row=2, column=COL_ANNUAL,
            value=f"=SUM({annual_letter}{first_data_row}:{annual_letter}{last_data_row})")
    style_total_row(ws, row=2, num_cols=NUM_COLS)

    # ── 5. Column headers (row 3) ────────────────────────────────────────
    ws.cell(row=3, column=COL_PANEL_ID, value="Panel ID")
    ws.cell(row=3, column=COL_MESH_ID,  value="Mesh ID")
    for m_i, month in enumerate(MONTHS):
        ws.cell(row=3, column=COL_JAN + m_i, value=month)
    ws.cell(row=3, column=COL_ANNUAL, value="Annual")
    style_column_headers(ws, row=3, num_cols=NUM_COLS, kpi="energy")

    # ── 6. Per-mesh data rows (row 4+) ───────────────────────────────────
    for i, (panel_id, mesh_id, monthly) in enumerate(rows):
        r = i + first_data_row
        ws.cell(row=r, column=COL_PANEL_ID, value=panel_id)
        ws.cell(row=r, column=COL_MESH_ID,  value=mesh_id)
        for m_i, val in enumerate(monthly):
            ws.cell(row=r, column=COL_JAN + m_i, value=val)
        ws.cell(row=r, column=COL_ANNUAL, value=_annual_formula(r))
        style_data_row(ws, row=r, num_cols=NUM_COLS, row_index=i)

    # ── 7. Column widths & freeze ─────────────────────────────────────────
    set_column_widths(ws, [12, 36] + [7] * len(MONTHS) + [12])
    ws.freeze_panes = f"{get_column_letter(COL_JAN)}4"