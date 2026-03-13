"""
excel_formatter.py
------------------
All Excel formatting helpers used across KPI sheets.

Colours use standard Excel theme hex values:
  - Light Magenta 2  : E4AFCA  (CFAR)
  - Light Red 2      : F4CCCC  (MUI)  
  - Light Yellow 2   : FFF2CC  (Energy)
  - Light Blue 2     : CFE2F3  (Modularity)
  - Light Grey 2     : EFEFEF  (alternating rows)
  - White            : FFFFFF  (alternating rows)

Font: Inter, black throughout.
All cells: center + middle aligned.
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from typing import List


# ─────────────────────────────────────────────────────────────
#  Colour palette
# ─────────────────────────────────────────────────────────────

class KPIColours:
    # KPI heading row backgrounds
    CFAR_HEADING        = "E4AFCA"   # Light Magenta 2
    MUI_HEADING         = "F4CCCC"   # Light Red 2
    ENERGY_HEADING      = "FFF2CC"   # Light Yellow 2
    MODULARITY_HEADING  = "CFE2F3"   # Light Blue 2

    # Column header row — slightly darker shade of each KPI colour
    CFAR_SUBHEADER      = "C9769E"
    MUI_SUBHEADER       = "E49090"
    ENERGY_SUBHEADER    = "FFD966"
    MODULARITY_SUBHEADER= "9FC5E8"

    # Data rows
    WHITE               = "FFFFFF"
    LIGHT_GREY_2        = "EFEFEF"

    # Universal
    BLACK               = "000000"


# ─────────────────────────────────────────────────────────────
#  Base styles
# ─────────────────────────────────────────────────────────────

ALIGN_CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Aliases used by kpi modules that import ALIGN and BORDER directly
ALIGN  = ALIGN_CENTER
BORDER = Border(
    left   = Side(style="thin", color="D0D0D0"),
    right  = Side(style="thin", color="D0D0D0"),
    top    = Side(style="thin", color="D0D0D0"),
    bottom = Side(style="thin", color="D0D0D0"),
)
THIN_BORDER = BORDER

# Shared fills and fonts for section subheaders (used by modularity)
FILL_SECTION   = PatternFill("solid", fgColor=KPIColours.MODULARITY_SUBHEADER)
FONT_SUBHEADER = Font(name="Inter", bold=True, color=KPIColours.BLACK, size=11)


def _font(bold: bool = False, size: int = 10) -> Font:
    return Font(name="Inter", bold=bold, color=KPIColours.BLACK, size=size)


def _fill(hex_colour: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_colour)


# ─────────────────────────────────────────────────────────────
#  Row styling functions
# ─────────────────────────────────────────────────────────────

def style_kpi_heading(ws, row: int, num_cols: int, kpi: str):
    """
    Style the KPI title row (e.g. 'Column Free Area Ratio').
    kpi must be one of: 'cfar', 'mui', 'energy', 'modularity'
    """
    colour_map = {
        "cfar":       KPIColours.CFAR_HEADING,
        "mui":        KPIColours.MUI_HEADING,
        "energy":     KPIColours.ENERGY_HEADING,
        "modularity": KPIColours.MODULARITY_HEADING,
    }
    fill_colour = colour_map.get(kpi.lower(), KPIColours.WHITE)

    ws.merge_cells(
        start_row=row, start_column=1,
        end_row=row, end_column=num_cols
    )
    cell = ws.cell(row=row, column=1)
    cell.font      = _font(bold=True, size=12)
    cell.fill      = _fill(fill_colour)
    cell.alignment = ALIGN_CENTER
    cell.border    = THIN_BORDER
    ws.row_dimensions[row].height = 28


def style_column_headers(ws, row: int, num_cols: int, kpi: str):
    """
    Style the column headers row (Level, Slab Area, etc.).
    Uses a darker shade of the KPI colour.
    """
    colour_map = {
        "cfar":       KPIColours.CFAR_SUBHEADER,
        "mui":        KPIColours.MUI_SUBHEADER,
        "energy":     KPIColours.ENERGY_SUBHEADER,
        "modularity": KPIColours.MODULARITY_SUBHEADER,
    }
    fill_colour = colour_map.get(kpi.lower(), KPIColours.LIGHT_GREY_2)

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = _font(bold=True, size=10)
        cell.fill      = _fill(fill_colour)
        cell.alignment = ALIGN_CENTER
        cell.border    = THIN_BORDER
    ws.row_dimensions[row].height = 22


def style_total_row(ws, row: int, num_cols: int):
    """
    Style the 'Entire Building' total row — bold, white background.
    """
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = _font(bold=True, size=10)
        cell.fill      = _fill(KPIColours.WHITE)
        cell.alignment = ALIGN_CENTER
        cell.border    = THIN_BORDER
    ws.row_dimensions[row].height = 20


def style_data_row(ws, row: int, num_cols: int, row_index: int):
    """
    Style a data row with alternating white / light grey.
    row_index: 0-based index of this data row (0 = first data row = white).
    """
    fill_colour = KPIColours.WHITE if row_index % 2 == 0 else KPIColours.LIGHT_GREY_2

    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = _font(bold=False, size=10)
        cell.fill      = _fill(fill_colour)
        cell.alignment = ALIGN_CENTER
        cell.border    = THIN_BORDER
    ws.row_dimensions[row].height = 18


# ─────────────────────────────────────────────────────────────
#  Sheet setup helpers
# ─────────────────────────────────────────────────────────────

def set_column_widths(ws, widths: List[int]):
    """Set column widths from a list, index 0 = column A."""
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def freeze_below_headers(ws, header_row: int):
    """Freeze all rows above and including the header row."""
    freeze_row = header_row + 1
    ws.freeze_panes = f"A{freeze_row}"