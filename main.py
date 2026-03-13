"""
main.py
-------
Speckle Automate function entry point.
"""

import openpyxl
from enum import Enum

from speckle_automate import (
    AutomateBase,
    AutomationContext,
    execute_automate_function,
)
from pydantic import Field

from kpi_cfar                  import write_cfar_sheet
from kpi_mui                   import write_mui_sheet
from kpi_modularity            import write_modularity_sheet
from kpi_energy                import write_energy_sheet
from transfer_modularity_model import transfer_modularity_model


class ISRMonth(str, Enum):
    january  = "January"
    february = "February"
    march    = "March"
    april    = "April"
    may      = "May"
    june     = "June"
    july     = "July"
    august   = "August"
    september= "September"
    october  = "October"
    november = "November"
    december = "December"
    annual   = "Annual"


class FunctionInputs(AutomateBase):
    modularity_model_stream_id: str = Field(
        title="Modularity Model ID",
        description="Model ID of the target model for the facade ISR transfer.",
    )
    modularity_model_branch: str = Field(
        default="main",
        title="Modularity Model Branch",
        description="Branch to commit the modularity model to.",
    )
    isr_month: ISRMonth = Field(
        default=ISRMonth.january,
        title="ISR Month",
        description="Month to use for the ISR heatmap colouring. Select 'Annual' for the sum of all 12 months.",
    )


def automate_function(
    automate_context: AutomationContext,
    function_inputs: FunctionInputs,
) -> None:

    version_root_object = automate_context.receive_version()

    if version_root_object is None:
        automate_context.mark_run_failed("Could not receive source model version.")
        return

    # ── KPI Excel ─────────────────────────────────────────────────────────
    try:
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        write_cfar_sheet(wb.create_sheet("CFAR"), version_root_object)
        write_mui_sheet(wb.create_sheet("MUI"), version_root_object)
        write_modularity_sheet(wb.create_sheet("Modularity Index"), version_root_object)
        write_energy_sheet(wb.create_sheet("Energy Performance"), version_root_object)
        excel_path = "/tmp/kpi_report.xlsx"
        wb.save(excel_path)
        automate_context.store_file_result(excel_path)
        print("KPI Excel report generated.")
    except Exception as e:
        automate_context.mark_run_failed(f"Excel generation failed: {e}")
        raise

    # ── Modularity Transfer ───────────────────────────────────────────────
    try:
        transfer_modularity_model(
            automate_context  = automate_context,
            speckle_client    = automate_context.speckle_client,
            version_root      = version_root_object,
            target_stream_id  = function_inputs.modularity_model_stream_id,
            target_branch     = function_inputs.modularity_model_branch,
            isr_month         = function_inputs.isr_month.value,
        )
        print("Modularity model transfer complete.")
    except Exception as e:
        automate_context.mark_run_failed(f"Modularity transfer failed: {e}")
        raise

    automate_context.mark_run_success(
        "KPI report generated and modularity model updated successfully."
    )


if __name__ == "__main__":
    execute_automate_function(automate_function, FunctionInputs)