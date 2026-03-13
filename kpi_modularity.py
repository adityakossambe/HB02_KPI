"""
kpi_modularity.py
-----------------
Populates the Modularity Index sheet in the KPI workbook.

Model structure:
    Grasshopper Model (root)
    ├── Facade       -> objects with property: panel_id  (e.g. T1, Q1)
    └── Exoskeleton  -> objects with property: member_id (e.g. LI1, NQ1)

Calculations:
    total_elements  = all meshes in Facade + Exoskeleton
    per unique ID:
        unit_count  = number of meshes sharing that ID
        normalised  = unit_count / total_elements
    building_index  = total_unique_ids / total_elements

Output layout:
    Row 1  — KPI heading (merged, Light Blue 2)
    Row 2  — Total Building Index
    Row 3  — Facade section subheader + column headers
    Row 4+ — Facade ID rows
    ...    — Exoskeleton section subheader + column headers
    ...    — Exoskeleton ID rows
"""

from collections import defaultdict
from openpyxl.styles import Font, PatternFill

from _collection_helper import get_collection_objects, get_prop
from excel_formatter import (
    style_kpi_heading,
    style_data_row,
    set_column_widths,
    freeze_below_headers,
    ALIGN,
    BORDER,
    FILL_SECTION,
    FONT_SUBHEADER,
)

NUM_COLS        = 3
SECTION_HEADERS = ["ID", "Unit Count", "Normalised Value (0-1)"]


def _apply(cell, fnt, fll):
    cell.font      = fnt
    cell.fill      = fll
    cell.alignment = ALIGN
    cell.border    = BORDER


def _count_ids(objects, id_key: str) -> dict:
    counts = defaultdict(int)
    for obj in objects:
        pid = get_prop(obj, id_key)
        if pid:
            counts[str(pid).strip()] += 1
    return dict(counts)


def _write_section(ws, start_row, section_title, id_counts, total_elements):
    """Write one collection section. Returns the next available row."""

    # Section subheader
    ws.merge_cells(
        start_row=start_row, start_column=1,
        end_row=start_row, end_column=NUM_COLS
    )
    _apply(ws.cell(row=start_row, column=1, value=section_title),
           FONT_SUBHEADER, FILL_SECTION)
    ws.row_dimensions[start_row].height = 24

    # Column headers
    header_row = start_row + 1
    for col_i, h in enumerate(SECTION_HEADERS, start=1):
        _apply(ws.cell(row=header_row, column=col_i, value=h),
               Font(name="Inter", bold=True, color="000000", size=10),
               FILL_SECTION)
    ws.row_dimensions[header_row].height = 20

    # Data rows
    for i, (uid, count) in enumerate(sorted(id_counts.items())):
        r = header_row + 1 + i
        normalised = count / total_elements if total_elements > 0 else 0.0
        ws.cell(row=r, column=1, value=uid)
        ws.cell(row=r, column=2, value=count)
        ws.cell(row=r, column=3, value=round(normalised, 4))
        style_data_row(ws, row=r, num_cols=NUM_COLS, row_index=i)
        ws.row_dimensions[r].height = 18

    return header_row + 1 + len(id_counts)


def write_modularity_sheet(ws, root):
    """Write the Modularity Index KPI data into worksheet ws."""

    # ── 1. Collect objects ───────────────────────────────────────────────
    facade_objects = get_collection_objects(root, "Facade")
    exo_objects    = get_collection_objects(root, "Exoskeleton")

    # ── 2. Count IDs ─────────────────────────────────────────────────────
    facade_id_counts = _count_ids(facade_objects, "panel_id")
    exo_id_counts    = _count_ids(exo_objects,    "member_id")

    total_elements   = len(facade_objects) + len(exo_objects)
    total_unique_ids = len(facade_id_counts) + len(exo_id_counts)
    building_index   = total_unique_ids / total_elements if total_elements > 0 else 0.0

    # ── 3. KPI heading (row 1) ───────────────────────────────────────────
    ws.cell(row=1, column=1, value="Modularity Index")
    style_kpi_heading(ws, row=1, num_cols=NUM_COLS, kpi="modularity")

    # ── 4. Total Building Index (row 2) ──────────────────────────────────
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=2)
    _apply(ws.cell(row=2, column=1, value="Total Building Index"),
           Font(name="Inter", bold=True, color="000000", size=10),
           PatternFill("solid", fgColor="FFFFFF"))
    _apply(ws.cell(row=2, column=3, value=round(building_index, 4)),
           Font(name="Inter", bold=True, color="000000", size=10),
           PatternFill("solid", fgColor="FFFFFF"))
    ws.row_dimensions[2].height = 22

    # ── 5. Facade section ────────────────────────────────────────────────
    next_row = _write_section(
        ws, start_row=3,
        section_title=f"Facade  ({len(facade_objects)} elements)",
        id_counts=facade_id_counts,
        total_elements=total_elements,
    )

    # ── 6. Exoskeleton section ───────────────────────────────────────────
    _write_section(
        ws, start_row=next_row + 1,
        section_title=f"Exoskeleton  ({len(exo_objects)} elements)",
        id_counts=exo_id_counts,
        total_elements=total_elements,
    )

    # ── 7. Column widths & freeze ─────────────────────────────────────────
    set_column_widths(ws, [16, 16, 26])
    freeze_below_headers(ws, header_row=2)