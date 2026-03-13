"""
excel_formatting.py
Applies colors and formatting to KPI Excel sheets.
"""

from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill

# Define heading colors
HEADING_COLORS = {
    "CFAR": "F4CCCC",        # Light Magenta 2
    "MUI": "F8CBAD",         # Light Red 2
    "Energy": "FFF2CC",      # Light Yellow 2
    "Modularity": "D9E1F2",  # Light Blue 2
}

# Alternate row colors
ROW_COLORS = ["FFFFFF", "D9D9D9"]  # White, Light Grey 2


def format_kpi_excel(filepath: str, kpi_name: str):
    """
    Apply formatting to a KPI Excel sheet.

    Args:
        filepath (str): Path to the Excel file.
        kpi_name (str): One of "CFAR", "MUI", "Energy", "Modularity".
    """
    wb = load_workbook(filepath)

    # Assume formatting is only needed for the first sheet
    ws = wb.active

    # Heading formatting
    heading_fill = PatternFill(start_color=HEADING_COLORS.get(kpi_name, "FFFFFF"), fill_type="solid")
    for cell in ws[1]:
        cell.fill = heading_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Data rows formatting
    for idx, row in enumerate(ws.iter_rows(min_row=2), start=0):
        fill_color = ROW_COLORS[idx % 2]
        for cell in row:
            cell.fill = PatternFill(start_color=fill_color, fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")

    wb.save(filepath)